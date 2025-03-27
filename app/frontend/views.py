from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.forms import ValidationError
from django.shortcuts import  render
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth.forms import PasswordChangeForm
import csv
import json
import openpyxl
from openpyxl import Workbook 
from django.contrib.auth import get_user_model
from django.views import View
from django.utils.text import slugify
from django.utils.timezone import make_naive
from datetime import datetime, timedelta

#local import
from django.shortcuts import render, redirect, get_object_or_404
from app.frontend.models import Customer
from app.frontend.forms import CustomerForm, ItemForm
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from app.frontend.models import Purchase, PurchaseItem
from app.frontend.models import Supplier
from app.frontend.models import Item, Category
from app.frontend.forms import PurchaseForm, PurchaseItemForm
from app.frontend.models import Category
from app.frontend.forms import CategoryForm
from app.frontend.models import Expense, ExpenseCategory
from app.frontend.forms import ExpenseForm, ExpenseCategoryForm,OrganizationSettingForm
from app.frontend.models import OrganizationSetting, Invoice
from app.frontend.models import Sale,SaleItem
from django.core.paginator import Paginator
from app.account.forms import StaffCreationForm




def index(request):
    return render(request, 'backend/index.html')


#For User/admin application like login, logout and change password
def login_view(request):
    try:
        if request.user.is_authenticated:
            return redirect('frontend:index')  

        if request.method == "POST":
            email = request.POST.get('email') 
            password = request.POST.get('password')

           
            user_obj = authenticate(request, username=email, password=password) 

            if user_obj is None:
                messages.warning(request, "Invalid email or password.")
                return redirect('frontend:login')  

            
            if user_obj.is_superuser or user_obj.is_staff or user_obj.is_editor:
                auth_login(request, user_obj)
                return redirect('frontend:index')  

            messages.warning(request, "Invalid credentials or insufficient permissions.")
            return redirect('frontend:login')

        return render(request, 'backend/login.html')

    except Exception as e:
        print(f"Login error: {e}") 
        messages.warning(request, "Something went wrong. Please try again.")
        return redirect('frontend:login')


@login_required
def userlogout(request):
    auth_logout(request)
    messages.info(request, "Logged out successfully.")
    return redirect('frontend:login')

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, 'Your password was successfully updated!')
            return redirect('frontend:change_password') 
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PasswordChangeForm(request.user)
    
    return render(request, 'backend/change_password.html', {'form': form})


def index(request):
    total_customers = Customer.objects.count() 
    total_items = Item.objects.count() 
    total_sale = Sale.objects.count() 
    total_category = Category.objects.count() 
    total_expenses = Expense.objects.count() 
    total_expenses_category= ExpenseCategory.objects.count() 
    return render(request, 'index.html',{'total_customers':total_customers,
    'total_items':total_items,
    'total_sale':total_sale,
    'total_category':total_category,
    'total_expenses':total_expenses,
    'total_expenses_category':total_expenses_category})


# Customer Details
def customer_list(request):
    customers = Customer.objects.all()
    
    paginator = Paginator(customers, 15)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'customer/customer_list.html', {
        'page_obj': page_obj,
        'customers': page_obj,  
        'is_paginated': page_obj.has_other_pages(),
    })



@login_required
def add_customer(request):
    if request.method == "POST":
        form = CustomerForm(request.POST)
        if form.is_valid():
            customer = form.save(commit=False)
            customer.created_by = request.user
            customer.save()
            messages.success(request, "Customer added successfully!")
            return redirect('frontend:customers_list')
    else:
        form = CustomerForm()
    return render(request, 'customer/add_customer.html', {'form': form})


@login_required
def delete_customer(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    customer.delete()
    messages.success(request, "Customer deleted successfully!")
    return redirect('frontend:customers_list')  



@login_required
def edit_customer(request, pk):
    customer = get_object_or_404(Customer, pk=pk)

    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, "Customer updated successfully!")
            return redirect('frontend:customers_list')
    else:
        form = CustomerForm(instance=customer)

    return render(request, 'customer/edit_customer.html', {'form': form, 'customer': customer})





def export_customers_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Customers"

    # Headers
    sheet.append(['ID', 'Created By', 'Customer Name', 'Email', 'Phone Number', 'Shipping Address'])

    # Data
    customers = Customer.objects.all()
    for customer in customers:
        sheet.append([
            customer.id,
            str(customer.created_by),  # Convert User object to string
            customer.name,
            customer.email,
            str(customer.phone_number),  # Convert PhoneNumber object to string
            customer.shipping_address
        ])

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="customers.xlsx"'
    workbook.save(response)

    return response



# Purchases Details

def purchase_list(request):
    purchases = Purchase.objects.all()
    return render(request, 'purchase/purchase_list.html', {'purchases': purchases})

def add_purchase(request):
    if request.method == 'POST':
        form = PurchaseForm(request.POST)
        if form.is_valid():
            purchase = form.save(commit=False)
            purchase.created_by = request.user  # Assign current user
            purchase.save()
            messages.success(request, 'Purchase added successfully!')
            return redirect('purchase_list')  # Redirect to list page
        else:
            messages.error(request, 'Error adding purchase. Please check the details.')
    else:
        form = PurchaseForm()
    
    suppliers = Supplier.objects.all()
    return render(request, 'purchase/add_purchase.html', {'form': form, 'suppliers': suppliers})

def edit_purchase(request, pk):
    purchase = get_object_or_404(Purchase, pk=pk)
    if request.method == 'POST':
        form = PurchaseForm(request.POST, instance=purchase)
        if form.is_valid():
            form.save()
            messages.success(request, 'Purchase updated successfully!')
            return redirect('purchase_list')
        else:
            messages.error(request, 'Error updating purchase.')
    else:
        form = PurchaseForm(instance=purchase)
    
    return render(request, 'purchase/edit_purchase.html', {'form': form, 'purchase': purchase})

def delete_purchase(request, pk):
    purchase = get_object_or_404(Purchase, pk=pk)
    if request.method == 'POST':
        purchase.delete()
        messages.success(request, 'Purchase deleted successfully!')
        return redirect('purchase_list')
    
    return render(request, 'purchases/delete_purchase.html', {'purchase': purchase})




#Items Details

def item_list(request):
    items_list = Item.objects.all()
    paginator = Paginator(items_list, 10)  # Show 10 items per page

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'items/item_list.html', {'page_obj': page_obj, 'is_paginated': page_obj.has_other_pages()})


def add_item(request):
    if request.method == "POST":
        form = ItemForm(request.POST, request.FILES)
        if form.is_valid():
            item = form.save(commit=False)
            item.created_by = request.user
            item.save()
            messages.success(request, "Item added successfully!")
            return redirect('frontend:item_list')
    else:
        form = ItemForm()
    return render(request, 'items/item_form.html', {'form': form})


def edit_item(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    if request.method == "POST":
        form = ItemForm(request.POST, request.FILES, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, "Item updated successfully!")
            return redirect('frontend:item_list')
    else:
        form = ItemForm(instance=item)
    return render(request, 'items/item_form.html', {'form': form})


def delete_item(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    item.delete()
    messages.success(request, "Item deleted successfully!")
    return redirect('frontend:item_list')



# Category  Details
def category_list(request):
    categories = Category.objects.all().order_by('-id')
    paginator = Paginator(categories, 10)  # Show 10 categories per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'category/category_list.html', {'page_obj': page_obj})

def category_create(request):
    if request.method == "POST":
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Category added successfully!")
            return redirect('frontend:category_list')
    else:
        form = CategoryForm()
    
    return render(request, 'category/category_form.html', {'form': form})

def category_update(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == "POST":
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, "Category updated successfully!")
            return redirect('frontend:category_list')
    else:
        form = CategoryForm(instance=category)
    
    return render(request, 'category/category_form.html', {'form': form, 'category': category})

def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    category.delete()
    messages.success(request, "Category deleted successfully!")
    return redirect('frontend:category_list')



# Pos Detail
def pos_page(request):
    categories = Category.objects.all()
    brands = Item.objects.values_list('brand', flat=True).distinct()
    items = Item.objects.all()
    invoice = Invoice.objects.create() 
    context = {
        'categories': categories,
        'brands': brands,
        'items': items,
        'invoice': invoice
    }
    return render(request, 'pos/pos_system.html', context)


# Generate Items COde
def generate_item_code(request):
    """Generate item_code dynamically based on name input."""
    name = request.GET.get('name', '')

    if not name.strip():
        return JsonResponse({'item_code': ''})  # Return empty if no name is provided

    initials = ''.join([word[0] for word in name.split()]).upper()  # Get initials
    last_item = Item.objects.filter(item_code__startswith=initials).order_by('-id').first()

    if last_item and last_item.item_code[len(initials):].isdigit():
        new_number = int(last_item.item_code[len(initials):]) + 1
    else:
        new_number = 1000  # Start numbering from 1000

    item_code = f"{initials}{new_number}"
    return JsonResponse({'item_code': item_code})



# Expenses Details
def expense_list(request):
    expenses = Expense.objects.all().order_by('-date')
    print(expenses)  
    paginator = Paginator(expenses, 10)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'expenses/expense_list.html', {'page_obj': page_obj})


# Create a new expense
def expense_create(request):
    if request.method == "POST":
        form = ExpenseForm(request.POST, request.FILES)  
        if form.is_valid():
            form.save()
            messages.success(request, "Expense added successfully!")
            return redirect('frontend:expense_list')
    else:
        form = ExpenseForm()
    return render(request, 'expenses/expense_form.html', {'form': form})

# Update an existing expense
def expense_update(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    if request.method == "POST":
        form = ExpenseForm(request.POST, instance=expense)
        if form.is_valid():
            form.save()
            messages.success(request, "Expense updated successfully!")
            return redirect('frontend:expense_list')
    else:
        form = ExpenseForm(instance=expense)
    return render(request, 'expenses/expense_form.html', {'form': form, 'expense': expense})

# Delete an expense
def expense_delete(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    expense.delete()
    messages.success(request, "Expense deleted successfully!")
    return redirect('frontend:expense_list')




# Expense Category 
@login_required
def expense_category_list(request):
    categories = ExpenseCategory.objects.all().order_by('-created_at')
    return render(request, 'expenses/category_list.html', {'categories': categories})

@login_required
def expense_category_add(request):
    if request.method == 'POST':
        form = ExpenseCategoryForm(request.POST)
        if form.is_valid():
            expense_category = form.save(commit=False)
            expense_category.created_by = request.user
            expense_category.save()
            messages.success(request, "Expense category added successfully!")
            return redirect('frontend:expense_category_list')
    else:
        form = ExpenseCategoryForm()
    return render(request, 'expenses/category_form.html', {'form': form, 'title': 'Add Expense Category'})

@login_required
def expense_category_edit(request, pk):
    category = get_object_or_404(ExpenseCategory, pk=pk)
    if request.method == 'POST':
        form = ExpenseCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, "Expense category updated successfully!")
            return redirect('frontend:expense_category_list')
    else:
        form = ExpenseCategoryForm(instance=category)
    return render(request, 'expenses/category_form.html', {'form': form, 'title': 'Edit Expense Category'})

@login_required
def expense_category_delete(request, pk):
    category = get_object_or_404(ExpenseCategory, pk=pk)
    category.delete()
    messages.success(request, "Expense category deleted successfully!")
    return redirect('frontend:expense_category_list')


# Search Customer Detail
@login_required
def search_customer(request):
    """Search for customers by name or phone number."""
    query = request.GET.get("query", "")
    customers = Customer.objects.filter(name__icontains=query)[:10]  # Limit results
    data = [{"id": c.id, "name": c.name, "phone": str(c.phone_number)} for c in customers]
    return JsonResponse(data, safe=False)



@login_required
def create_customer(request):
    """Handle customer creation via AJAX."""
    if request.method == "POST":
        form = CustomerForm(request.POST)
        if form.is_valid():
            customer = form.save(commit=False)
            customer.created_by = request.user
            customer.save()
            return JsonResponse({"success": True, "id": customer.id, "name": customer.name})
        else:
            return JsonResponse({"success": False, "errors": form.errors})
    return JsonResponse({"success": False, "error": "Invalid request"})


# Organization Detail 
def organization_setting_view(request):
    # Get the first OrganizationSetting instance (assuming only one entry is allowed)
    organization = OrganizationSetting.objects.first()

    if request.method == 'POST':
        form = OrganizationSettingForm(request.POST, request.FILES, instance=organization)
        if form.is_valid():
            form.save()
            messages.success(request, "Organization settings updated successfully!")
            return redirect('frontend:organization_setting')  # Redirect to refresh the page
    else:
        form = OrganizationSettingForm(instance=organization)

    return render(request, 'organization_setting.html', {'form': form})






from django.shortcuts import render
from django.core.paginator import Paginator
from django.db.models import Q
from .models import Sale
from datetime import datetime


from datetime import datetime
from django.utils.timezone import make_aware

def sale_list(request):
    # Get the date range from the request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Start with all sales
    sales_list = Sale.objects.all().order_by('-created_at')

    # Filter by date range if provided
    if start_date and end_date:
        try:
            # Convert the date strings to datetime objects
            start_date = make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            end_date = make_aware(datetime.strptime(end_date, '%Y-%m-%d'))
            
            # Include the entire end date by adding 1 day
            end_date = end_date + timedelta(days=1)
            
            # Filter sales between the dates
            sales_list = sales_list.filter(created_at__range=(start_date, end_date))
        except ValueError:
            # Handle invalid date format - will return all sales
            pass
    elif start_date:
        try:
            # Filter sales after start date
            start_date = make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            sales_list = sales_list.filter(created_at__gte=start_date)
        except ValueError:
            pass
    elif end_date:
        try:
            # Filter sales before end date (including the entire end date)
            end_date = make_aware(datetime.strptime(end_date, '%Y-%m-%d')) + timedelta(days=1)
            sales_list = sales_list.filter(created_at__lte=end_date)
        except ValueError:
            pass

    # Paginate the results
    paginator = Paginator(sales_list, 10)  # Show 10 sales per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'sale/sale_list.html', {
        'sales': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
    })

def sale_detail(request, sale_id):
    sale = get_object_or_404(Sale, id=sale_id)
    sale_items = sale.saleitem_set.all()  # Fetch related items
    print("Sale:", sale)  # Debugging statement
    print("Sale Items:", sale_items)  # Debugging statement
    return render(request, 'sale/sale_detail.html', {'sale': sale, 'sale_items': sale_items})


# Save Sale List
def save_sale(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        print("Received data:", data)  # Debugging statement

        customer_name = data.get('customerName')
        customer_address = data.get('customerAddress')
        customer_phone = data.get('customerPhone')
        items = data.get('items')
        total_amount = data.get('totalAmount')

        print("Items:", items)  # Debugging statement

        # Create the Sale object
        sale = Sale.objects.create(
            customer_name=customer_name,
            customer_address=customer_address,
            customer_phone=customer_phone,
            total_amount=total_amount
        )

        # Create SaleItem objects for each item in the sale
        for item in items:
            try:
                item_obj = Item.objects.get(id=item['itemId'])  # Ensure the item exists
                print("Found Item:", item_obj.name)

                SaleItem.objects.create(
                    sale=sale,
                    item=item_obj,
                    item_name=item_obj.name,  # Store the item name
                    price=item_obj.price,  # Store the price at the time of sale
                    quantity=item['quantity'],
                    discount=item['discount'],
                    tax=item['tax'],
                    subtotal=item['subtotal']  # Store the subtotal
                )
            except Item.DoesNotExist:
                print(f"Item with ID {item['itemId']} not found.")

        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})





User = get_user_model()

class StaffListView(View):
    template_name = 'backend/staff_list.html'

    def get(self, request):
        staff_members = User.objects.filter(is_staff=True)
        return render(request, self.template_name, {'staff_members': staff_members})



    
def delete_staff(request, pk):
    staff_member = get_object_or_404(User, pk=pk)
    staff_member.delete()
    messages.success(request, "Staff member deleted successfully!")
    return redirect('frontend:staff_list')  

class StaffCreateView(View):
    template_name = 'backend/create_staff.html'

    def get(self, request):
        form = StaffCreationForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = StaffCreationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password'])  # Hash password
            user.is_staff = True  # Mark as staff

            # Generate a unique username from email
            base_username = slugify(user.email.split('@')[0])  # Use email prefix
            counter = 1
            username = base_username
            while User.objects.filter(username=username).exists():
                username = f"{base_username}{counter}"
                counter += 1

            user.username = username  # Assign generated username
            user.save()

            messages.success(request, "Staff member created successfully!")
            return redirect('frontend:staff_list')

        return render(request, self.template_name, {'form': form}) 


def export_sales_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sales"

    # Headers
    sheet.append(['Sale ID', 'Customer Name', 'Customer Phone', 'Total Amount', 'Date'])

    # Fetch sales data
    sales = Sale.objects.all()
    for sale in sales:
        # Convert datetime to naive (remove timezone)
        created_at_naive = make_naive(sale.created_at)

        sheet.append([
            sale.id, 
            sale.customer_name, 
            sale.customer_phone, 
            sale.total_amount, 
            created_at_naive  # Ensure this is a naive datetime
        ])

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sales.xlsx"'
    workbook.save(response)

    return response


def export_items_excel(request):
    # Create an Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Items List"

    # Define the header row
    headers = ["ID", "Item Code", "Name", "Brand", "Category", "Quantity", "Price"]
    sheet.append(headers)

    # Fetch items from the database
    items = Item.objects.all()

    # Write item data to the Excel file
    for item in items:
        sheet.append([item.id, item.item_code, item.name, item.brand, item.category, item.quantity, item.price])

    # Prepare the HTTP response with the Excel file
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="items_list.xlsx"'
    workbook.save(response)

    return response